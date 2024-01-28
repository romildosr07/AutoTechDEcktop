from tkinter import END

import openpyxl
from datetime import datetime

from AutoTechDecktop.database import PartControl, ServiceControl, ServiceOrderControl, ReportControl


class MakeOrder:
    def make_excel(self):

        service_oreder_control = ServiceOrderControl('basededados.db')
        list_os = service_oreder_control.read_all_service_order()

        vehicle = ""
        plate = ""
        kms = ""
        for os in list_os:
            vehicle = os[1]
            plate = os[2]
            kms = os[3]

        book = openpyxl.load_workbook('static/autotech.xlsx')
        sheet = book['OS']

        book2 = openpyxl.load_workbook('static/base_excel.xlsx')
        base_excel = book2['base']

        # numero da os
        numero = sheet['N5'].value
        sheet['N5'] = numero + 1

        book.save('static/autotech.xlsx')

        book = openpyxl.load_workbook('static/autotech.xlsx')
        sheet = book['OS']

        numero_os = sheet['N5'].value


        sheet['G9'] = vehicle
        sheet['I9'] = plate.upper()
        sheet['L9'] = kms



        cell_number = 15

        cell_qtd = 'C15'
        cell_part = 'D15'
        cell_value = 'N15'

        part_control = PartControl('basededados.db')
        part_list = part_control.read_all_part()

        total_part = 0
        for part in part_list:
            sheet[cell_qtd] = part[1]
            sheet[cell_part] = part[2]
            sheet[cell_value] = part[3]

            cell_number += 1

            cell_qtd = f'C{cell_number}'
            cell_part = f'D{cell_number}'
            cell_value = f'N{cell_number}'

            total_part += part[3]

            if cell_number == 37:
                break



        # Service
        service_control = ServiceControl('basededados.db')
        service_list = service_control.read_all_service()

        cell_number = 42

        cell_service = 'C42'
        cell_value = 'N42'

        total_service = 0
        for service in service_list:
            sheet[cell_service] = service[1]
            sheet[cell_value] = service[2]

            cell_number += 1

            cell_service = f'C{cell_number}'
            cell_value = f'N{cell_number}'

            total_service += service[2]
            if cell_number == 53:
                break

        data = datetime.now()

        name_os = f'ordem_de_serviço/{plate}-{data.strftime("%d-%m-%y")}.xlsx'
        book.save(name_os)

        total_general = total_service + total_part

        med = Report()
        med.report(numero_os, plate, vehicle, kms, data.strftime("%d/%m/%y %H:%m"), total_part, total_service, total_general)


class Report:
    def report(self, numero_os, plate, vehicle, kms, date, total_part, total_service, total_general):
        report_control = ReportControl('basededados.db')
        report_control.create_report(numero_os, plate, vehicle, kms, date, total_part, total_service, total_general)



